import json
import os
import random
import re
import shutil
import sys
import time
import serial
import requests
from zipfile import ZipFile
from utils.functions.gpio import GPIO
from utils.operate.at_handle import ATHandle
from utils.logger.logging_handles import all_logger
from utils.functions.jenkins import multi_thread_merge_version
from utils.functions.driver_check import DriverChecker
from utils.exception.exceptions import LinuxCloudOTAError
from openpyxl import Workbook


class LinuxCloudOTAManager:
    def __init__(self, at_port, dm_port, prev_firmware_path, firmware_path, module_model, prev_ati, ati, prev_sub, sub, imei, debug_port):
        self.at_port = at_port
        self.debug_port = debug_port
        self.dm_port = dm_port
        self.module_model = module_model    # 硬件型号，例RG500Q-EA,用来创建模组类型
        self.at_handle = ATHandle(at_port)
        self.driver_check = DriverChecker(at_port, dm_port)
        self.gpio = GPIO()
        self.prev_firmware_path = prev_firmware_path    # 上一版本升级包共享路径
        self.firmware_path = firmware_path  # 当前版本升级包共享路径
        self.prev_ati = prev_ati    # 上一版本的ATI号
        self.prev_sub = prev_sub    # 上一版本的SUB号
        self.ati = ati    # 当前版本的ATI号
        self.sub = sub    # 当前版本的SUB号
        self.imei = imei
        self.ori_authorization = self.get_token()
        self.authorization = {'authorization': f'Bearer {self.ori_authorization}'}   # 获取token，之后所有的请求都带token请求
        self.user_id = self.get_user_id()   # 获取user_id,用于部分请求拼接URL
        self.project_id = self.get_projectId()  # 获取project_id,用于部分请求拼接URL，例如QP0001L3
        project_info = self.get_product_id_key()     # 获取product_id及product_id
        self.product_id = project_info[0]  # 获取product_id,用于发请求， 例如206
        self.project_key = project_info[1]  # 获取project_key,用于发请求， 例如p111SQ
        all_logger.info(f"authorization:'{self.authorization}',user_id: '{self.user_id}'"
                        f",project_id: '{self.project_id}', product_id: '{self.product_id}'")
        self.version_flag = self.check_revision()    # 确定当前版本号，用于后续差分升级后版本检查

    def check_merge_version(self):
        """
        检查之前是否执行过case且已制作完成差分包，有的话直接复制到当前执行文件夹下
        :return:
        """
        package_name = self.firmware_path.split('\\')[-1]
        prev_package_name = self.prev_firmware_path.split('\\')[-1]
        all_logger.info(f'当前版本包名为:{package_name}，上一版本包名为{prev_package_name}')
        path = '/root/TWS_TEST_DATA'
        package_flag = False
        prev_package_flag = False
        package_path = ''
        for i in os.listdir(path):
            if 'CloudOTA' in i:
                for path, dirs, files in os.walk(os.path.join(path, i)):
                    if package_name in dirs:
                        package_flag = True
                        package_path = path
                    if prev_package_name in dirs:
                        prev_package_flag = True
                    if package_flag and prev_package_flag:
                        if 'a-b.zip' in os.listdir(os.path.abspath(os.path.dirname(package_path))) and 'b-a.zip' in os.listdir(os.path.abspath(os.path.dirname(package_path))):
                            all_logger.info('之前存在已做好的差分包，复制到当前文件夹下后使用')
                            tar = os.path.abspath(os.path.dirname(path))
                            os.mkdir(os.path.join(os.getcwd(), 'firmware'))
                            shutil.copy(os.path.join(tar, 'a-b.zip'), os.path.join(os.getcwd(), 'firmware', 'a-b.zip'))
                            shutil.copy(os.path.join(tar, 'b-a.zip'), os.path.join(os.getcwd(), 'firmware', 'b-a.zip'))
                            all_logger.info(f'当前文件夹下文件有{os.listdir(os.path.join(os.getcwd(), "firmware"))}')
                            return True
        else:
            return False

    def check_revision(self):
        """
        ATI检查当前版本,如果当前版本为测试版本，则下一次差分升级后为其他版本，若当前为其他版本，则相反
        :return:
        """
        ati_value = self.at_handle.send_at('ATI')
        reversion = ''.join(re.findall(r'Revision: (.*)', ati_value)).strip()
        if reversion == self.ati:   # 如果是当前测试版本则返回True
            all_logger.info('当前版本为测试版本')
            return True
        else:
            all_logger.info('当前版本为差分升级后的目标版本')
            return False

    def query_default_fota_value(self):
        """
        检测AT+QFOTACFG及AT+QFOTAUP指令默认值
        :return:
        """
        cfg_value_1 = self.at_handle.send_at('AT+QFOTACFG=?', 10)
        cfg_value_2 = self.at_handle.send_at('AT+QFOTACFG?', 10)
        up_value_1 = self.at_handle.send_at('AT+QFOTAUP=?', 10)
        up_value_2 = self.at_handle.send_at('AT+QFOTAUP?', 10)
        cfg_default_value_list = ['+QFOTACFG: "server",<url>', '+QFOTACFG: "tls",(0,1)', '+QFOTACFG: "pk",<secret>,<key>']
        for i in cfg_default_value_list:
            if i not in cfg_value_1:
                raise LinuxCloudOTAError(f'AT+QFOTACFG=?默认值不正确，缺少{i}返回值,返回值为{cfg_value_1}')
        if '+QFOTACFG: "tls",0' not in cfg_value_2:
            raise LinuxCloudOTAError(f'AT+QFOTACFG?返回值不正确，+QFOTACFG: "tls"返回值不为0, 返回值为{cfg_value_2}')
        if '+QFOTAUP: (0-100),(1,2),<url>' not in up_value_1:
            raise LinuxCloudOTAError(f'AT+QFOTAUP=?返回值不正确，返回值为{up_value_1}')
        if '+QFOTAUP: 1' not in up_value_2:
            raise LinuxCloudOTAError(f'AT+QFOTAUP?返回值不正确，<upmode>参数不为1，返回值为{up_value_2}')
        all_logger.info('AT+QFOTACFG及AT+QFOTAUP指令默认值检查成功')

    @staticmethod
    def make_dfota_package():
        """
        查找当前路径 + firmware + prev/cur路径下面的所有targetfiles.zip，然后制作差分包。
        :return: None
        """
        all_logger.info("开始制作差分包")
        orig_target_file = ''
        cur_target_file = ''

        firmware_path = os.path.join(os.getcwd(), 'firmware')
        for path, _, files in os.walk(os.path.join(firmware_path, 'prev')):
            for file in files:
                if file == 'targetfiles.zip':
                    orig_target_file = os.path.join(path, file)
        if not orig_target_file:
            raise LinuxCloudOTAError("获取前一个版本target file zip失败")

        for path, _, files in os.walk(os.path.join(firmware_path, 'cur')):
            for file in files:
                if file == 'targetfiles.zip':
                    cur_target_file = os.path.join(path, file)
        if not cur_target_file:
            raise LinuxCloudOTAError("获取当前版本target file zip失败")

        multi_thread_merge_version(orig_target_file, cur_target_file)

    def mount_package(self):
        """
        挂载版本包
        :return:
        """
        cur_package_path = self.firmware_path.replace('\\\\', '//').replace('\\', '/').replace(' ', '\\ ')
        prev_package_path = self.prev_firmware_path.replace('\\\\', '//').replace('\\', '/').replace(' ', '\\ ')
        if 'cur' not in os.listdir('/mnt'):
            os.mkdir('/mnt' + '/cur')
        if 'prev' not in os.listdir('/mnt'):
            os.mkdir('/mnt' + '/prev')
        os.popen('mount -t cifs {} /mnt/cur -o user="cris.hu@quectel.com",password="hxc111...."'.format(cur_package_path)).read()
        os.popen('mount -t cifs {} /mnt/prev -o user="cris.hu@quectel.com",password="hxc111...."'.format(prev_package_path)).read()
        if os.listdir('/mnt/cur') and os.listdir('/mnt/prev'):
            all_logger.info('版本包挂载成功')
        else:
            all_logger.info('版本包挂载失败,请检查版本包路径是否正确，或路径下是否存在版本包')
            raise LinuxCloudOTAError('版本包挂载失败,请检查版本包路径是否正确，或路径下是否存在版本包')

    @staticmethod
    def umount_package():
        """
        卸载已挂载的内容
        :return:
        """
        os.popen('umount /mnt/cur')
        os.popen('umount /mnt/prev')

    def ubuntu_copy_file(self):
        """
        Ubuntu下复制版本包
        :return:
        """
        os.mkdir(os.getcwd() + '/firmware')
        os.mkdir(os.path.join(os.getcwd(), 'firmware') + '/cur')
        os.mkdir(os.path.join(os.getcwd(), 'firmware') + '/prev')

        cur_file_list = os.listdir('/mnt/cur')
        all_logger.info('/mnt/cur目录下现有如下文件:{}'.format(cur_file_list))
        all_logger.info('开始复制当前版本版本包到本地')
        for i in cur_file_list:
            if self.ati.upper() in i.upper() and 'factory' not in i:   # 过滤掉工厂版本及其他文件，只复制标准包
                shutil.copy(os.path.join('/mnt/cur', i), os.path.join(os.getcwd(), 'firmware', 'cur'))

        prev_file_list = os.listdir('/mnt/prev')
        all_logger.info('/mnt/prev目录下现有如下文件:{}'.format(prev_file_list))
        all_logger.info('开始复制上一版本版本包到本地')
        for i in prev_file_list:
            if self.prev_ati.upper() in i.upper() and 'factory' not in i:   # 过滤掉工厂版本及其他文件，只复制标准包
                shutil.copy(os.path.join('/mnt/prev', i), os.path.join(os.getcwd(), 'firmware', 'prev'))

        if os.path.join(os.getcwd(), 'firmware', 'cur') and os.path.join(os.getcwd(), 'firmware', 'prev'):
            all_logger.info('版本获取成功')
        else:
            raise LinuxCloudOTAError('版本包获取失败')

    @staticmethod
    def unzip_firmware():
        """
        解压当前路径 + firmware + prev/cur路径下面的所有zip包
        :return: None
        """
        firmware_path = os.path.join(os.getcwd(), 'firmware')
        for path, _, files in os.walk(firmware_path):
            for file in files:
                if file.endswith('.zip'):
                    with ZipFile(os.path.join(path, file), 'r') as to_unzip:
                        all_logger.info('exact {} to {}'.format(os.path.join(path, file), path))
                        to_unzip.extractall(path)
        all_logger.info('解压固件成功')

    @staticmethod
    def get_token():
        """
        获取OTA平台token
        :return:
        """
        login_url = 'https://iot-gateway.quectel.com/v2/cloudotaportal/user/login?username=ruby&password=c09d68842844d1b989126232c67adbd0'
        login_req = requests.post(login_url).json()
        if not login_req['success']:
            all_logger.info(f'请求获取token接口返回:{login_req}')
        return login_req['result']   # 首先用账户密码登录获取token，之后的所有请求都带token请求

    def get_user_id(self):
        """
        获取当前登录用户的userid，用于拼接URL
        :return:
        """
        authorization = self.authorization['authorization'][7:]     # 获取的token需要处理回原始样子
        user_id_url = f'https://iot-gateway.quectel.com/v2/cloudotaportal/user/info?token={authorization}'
        user_id_req = requests.get(url=user_id_url, headers=self.authorization).json()
        if not user_id_req['success']:
            all_logger.info(f'请求获取用户user_id接口返回：{user_id_req}')
        user_id = user_id_req['result']['id']
        return user_id

    def get_projectId(self):
        """
        获取项目列表页中项目的project_id
        :return:
        """
        cur_project_url = f'https://iot-gateway.quectel.com/v2/cloudotaportal/cloudproject/list?rows=10&page=1&userId={self.user_id}&projectName='
        cur_project_req = requests.get(url=cur_project_url, headers=self.authorization).json()
        if not cur_project_req['success']:
            all_logger.info(f'请求获取项目列表信息接口返回：{cur_project_req}')
        cur_project_name = cur_project_req['result']['content']
        all_logger.info(f'现有的项目信息为:{cur_project_name}')
        for i in cur_project_name:
            if self.module_model in i['projectName']:
                return i['projectId']
        else:   # 当前系统中还不存在该型号的项目，需要新建，之后再重新重新返回projectId
            project_name = self.add_project()
            cur_project_req = requests.get(url=cur_project_url, headers=self.authorization).json()
            if not cur_project_req['success']:
                all_logger.info(f'请求获取项目列表信息接口返回：{cur_project_req}')
            cur_project_name = cur_project_req['result']['content']
            all_logger.info(f'现有的项目信息为:{cur_project_name}')
            for i in cur_project_name:
                if project_name in i['projectName']:
                    return i['projectId']

    def add_project(self):
        """
        项目列表页添加项目
        :return:
        """
        # 项目列表页面添加项目
        add_project_url = 'https://iot-gateway.quectel.com/v2/cloudotaportal/cloudproject/saveCloudProject'
        data = {'projectName': f'{self.module_model}_ota_upgrade',
                'remark': ''}
        add_pro_req = requests.post(url=add_project_url, headers=self.authorization, data=data)
        if not add_pro_req.json()['success']:
            all_logger.info(f'请求项目列表页添加项目接口返回：{add_pro_req.json()}')
        if add_pro_req.status_code == 200:
            all_logger.info(f'{self.module_model}_ota_upgrade')
            return f'{self.module_model}_ota_upgrade'

    def get_product_id_key(self):
        """
        获取产品列表页面，如果不存在当前测试版本，进行添加,之后返回产品id
        :return: 返回产品id及产品密钥，用于后续上传差分包URL拼接
        """
        product_url = f'https://iot-gateway.quectel.com/v2/cloudotaportal/project/list?page=1&rows=10&projectName=&userId={self.user_id}&cloudProjectId={self.project_id}'
        product_req = requests.get(url=product_url, headers=self.authorization).json()
        if not product_req['success']:
            all_logger.info(f'请求产品列表页面现有产品接口返回：{product_req}')
        cur_product_list = product_req['result']['content']
        for i in cur_product_list:
            if self.module_model in i['projectName']:
                all_logger.info('当前测试版本已存在于系统中，可直接上传差分包进行测试')
                return [i['id'], i['projectKey']]
        else:
            all_logger.info('当前测试版本尚未存在于系统，进行添加')
            module_id = self.get_module_id()    # 获取当前测试硬件型号在系统中的硬件ID
            add_product_url = f'https://iot-gateway.quectel.com/v2/cloudotaportal/project/save?projectName={self.module_model}_ota_upgrade&type=0&moduleId={module_id}&details=版本测试&isZh=true&qbProjectId={self.project_id}'
            add_product_req = requests.post(url=add_product_url, headers=self.authorization).json()
            if not add_product_req['success']:
                all_logger.info(f'请求产品列表页添加产品接口返回：{add_product_req}')
            if '成功' in add_product_req['msg']:
                all_logger.info('测试版本添加成功')

                # 添加成功后，再请求一次，获取产品id返回
                product_req = requests.get(url=product_url, headers=self.authorization).json()
                if not product_req['success']:
                    all_logger.info(f'请求产品列表页面现有产品接口返回：{product_req}')
                cur_product_list = product_req['result']['content']
                for i in cur_product_list:
                    if self.module_model in i['projectName']:
                        return [i['id'], i['projectKey']]

    def get_module_id(self):
        """
        获取模组对应的module_id，如果系统中不存在该模组，则创建并返回module_id
        :return:
        """
        module_png = {'RG50XQ': r'utils/images/module_pic/RG50XQ.png', 'RG500U-CN': r'utils/images/module_pic/RG50XQ.png',
                      'RM500Q-AE': r'utils/images/module_pic/RM500Q-AE.png', 'RM500Q-GL': r'utils/images/module_pic/RM500Q-GL.png',
                      'RM500U-CN': r'utils/images/module_pic/RM500U-CN.png', 'RM510Q': r'utils/images/module_pic/RM510Q.png',
                      'others': r'utils/images/module_pic/no_data.jpg'}    # 确定模块图片地址，用于添加模组上传
        module_id_url = 'https://iot-gateway.quectel.com/v2/cloudotaportal/deviceModel/getDeviceModelListForCustomer'
        add_module_url = 'https://iot-gateway.quectel.com/zuul/v2/cloudotaportal/deviceModel/save'     # 添加模组信息url
        id_req = requests.get(url=module_id_url, headers=self.authorization).json()
        if not id_req['success']:
            all_logger.info(f'请求获取当前模组列表接口返回：{id_req}')
        for i in id_req['result']:
            if i['moduleType'] == self.module_model:
                all_logger.info('当前已存在该模组类型')
                module_id = i['id']
                return module_id
        else:   # 如果没查到模组名，则新建模组
            all_logger.info('当前尚不存在该模组类型，需要添加')
            # 拼接图片地址
            module_model = self.module_model
            if 'RG50' in self.module_model:
                module_model = 'RG50XQ'
            try:
                module_png[module_model]
            except KeyError:
                module_model = 'others'
            all_logger.info('sys.path[0]: {}'.format(sys.path[0]))
            png_path = os.path.join(sys.path[0], module_png[module_model])
            all_logger.info(f'png_path: {png_path}')

            module_data = {'moduleType': f'{self.module_model}', 'details': 'Test', 'isCustomerOperation': True}
            files = {'file': open(png_path, 'rb')}

            # 请求添加模组
            add_module_req = requests.post(url=add_module_url, data=module_data, headers=self.authorization, files=files)
            all_logger.info(f'请求添加模组接口返回:{add_module_req.json()}')

            # 再获取module_id
            id_req = requests.get(url=module_id_url, headers=self.authorization).json()
            if not id_req['success']:
                all_logger.info(f'请求获取当前模组列表接口返回：{id_req}')
            for i in id_req['result']:
                if i['moduleType'] == self.module_model:
                    module_id = i['id']
                    return module_id

    def check_version(self, is_forward):
        """
        检查当前是否存在已发布的测试版本，不存在的话创建版本并发布
        :param is_forward:是否是正向升级：True:正向升级，False:反向升级
        :return:
        """
        # 从产品列表页点击详情进入产品信息页，再点击版本管理,查看当前是否存在测试版本，不存在则添加版本
        standard_url = f'https://iot-gateway.quectel.com/v2/cloudotaportal/version/listByPageForUser?page=1&size=50&projectId={self.product_id}'
        standard_info = requests.post(url=standard_url, headers=self.authorization).json()
        if not standard_info['success']:
            all_logger.info(f'请求产品列表页当前产品接口返回：{standard_info}')
        current_version = self.ati if is_forward else self.prev_ati
        target_version = self.prev_ati if is_forward else self.ati
        all_logger.info('上传正向差分包') if is_forward else all_logger.info('上传反向差分包')
        for i in standard_info['result']['data']:
            if i['currentVersion'] == current_version and i['targetVersion'] == target_version and i['releaseStatus'] == 'RELEASED':
                all_logger.info('当前版本管理页中已存在版本从当前版本差分升级到上一版本')
                return True
            elif i['releaseStatus'] != 'RELEASED':
                delete_id = i['id']
                # 如果不是且未发布就删除该版本
                delete_standard_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/version/delete?id={delete_id}'
                delete_req = requests.get(url=delete_standard_url, headers=self.authorization).json()
                if not delete_req['success']:
                    all_logger.info(f'请求删除无关版本接口返回：{delete_req}')
                else:
                    all_logger.info('删除无关版本')
        else:
            all_logger.info('当前版本管理页中不存在版本从当前版本差分升级到上一版本，需要新建版本')
            self.create_version(True) if is_forward else self.create_version(False)

    def create_version(self, is_forward):
        """
        添加版本并且上传差分包
        :param is_forward:是否是正向升级：True:正向升级，False:反向升级
        :return:
        """
        # 开始添加版本，第一步，首先创建信息
        forward_path = os.path.join(os.getcwd(), 'firmware', 'b-a.zip')     # 正向升级差分包路径，升级后不为测试版本
        reverse_path = os.path.join(os.getcwd(), 'firmware', 'a-b.zip')     # 反向升级差分包路径，升级后变为测试版本
        path = forward_path if is_forward else reverse_path
        package_name = path.split('\\')[-1]     # 切片提取差分包名用于传参
        package_size = os.path.getsize(path)    # 获取差分包大小用于传参
        create_url = 'https://iot-gateway.quectel.com/v2/cloudotaportal/version/createMultiPartUpload'  # 添加版本URL

        data = {"fileName": package_name, "projectId": self.project_id, "productKey": self.project_key,
                'totalSize': package_size}

        headers = {'authorization': f'Bearer {self.ori_authorization}', 'Content-Type': 'application/json'}
        create_req = requests.post(url=create_url, headers=headers, data=json.dumps(data)).json()
        if create_req['code'] != 200:
            all_logger.info(f'请求添加产品接口返回：{create_req}')
        upload_id = create_req['data']['uploadId']      # 获取该ID用于后续传参

        # 第二步，上传差分包

        upload_url = 'https://iot-gateway.quectel.com/zuul/v2/cloudotaportal/version/uploadMultipart'
        files = {'file': open(path, 'rb')}
        info = json.dumps({"uploadId": upload_id})
        data = {"chunkNumber": 1, "currentChunkSize": package_size, "totalSize": package_size,
                "identifier": info, "filename": package_name, "relativePath": package_name}
        upload_req = requests.post(url=upload_url, headers=self.authorization, data=data, files=files).json()
        if upload_req['code'] != '200':
            all_logger.info(upload_req)

        # 第三步，将初始版本，目标版本等信息附加带上进行请求
        complete_url = 'https://iot-gateway.quectel.com/v2/cloudotaportal/version/completeMultipart'

        data = {"projectId": f"{self.project_id}", "productId": f"{self.product_id}",
                "productKey": f"{self.project_key}", "packageType": "DIFF_PACK",
                "currentVersion": self.ati if is_forward else self.prev_ati, "targetVersion": self.prev_ati if is_forward else self.ati,
                "versionInfo": "test", "uploadId": f"{upload_id}", "signType": "md5",
                }
        complete_req = requests.post(url=complete_url, headers=self.authorization, data=data).json()
        if not complete_req['code'] != '200':
            all_logger.info(f'请求完成添加产品信息接口返回：{complete_req}')
        file_crc = complete_req['data']['fileCrc']
        file_md5 = complete_req['data']['fileMd5']
        file_sha256 = complete_req['data']['fileSha256']
        save_id = complete_req['data']['id']
        file_path = complete_req['data']['filePath']

        # 第四步，保存信息提交
        save_url = 'https://iot-gateway.quectel.com/v2/cloudotaportal/version/save'

        data = {"fileList": [{"id": save_id, "fileCrc": file_crc, "fileMd5": file_md5, "filePath": file_path,
                              "fileSize": package_size, "fileSha256": file_sha256}],
                "currentVersion": self.ati if is_forward else self.prev_ati, "targetVersion": self.prev_ati if is_forward else self.ati,
                "packageType": "DIFF_PACK", "projectId": self.product_id, "versionInfo": "test"}

        headers = {'authorization': f'Bearer {self.ori_authorization}', 'Content-Type': 'application/json'}
        save_req = requests.post(url=save_url, headers=headers, data=json.dumps(data)).json()
        if save_req['code'] != '200':
            all_logger.info(f'请求保存信息提交接口返回：{save_req}')

        # 第五步，发布版本
        # 首先查看已添加的版本，找到待发布版本的id，用于发布
        standard_manage_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/version/listByPageForUser?page=1&size=50&projectId={self.product_id}'
        cur_standard = requests.post(url=standard_manage_url, headers=self.authorization).json()['result']['data']
        standard_id = ''
        for i in cur_standard:
            cur_version = self.ati if is_forward else self.prev_ati
            tar_version = self.prev_ati if is_forward else self.ati
            if cur_version == i['currentVersion'] and tar_version == i['targetVersion']:
                standard_id = i['id']

        release_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/version/release?id={standard_id}'
        release_req = requests.get(url=release_url, headers=self.authorization).json()
        if not release_req['success']:
            all_logger.info(f'请求发布版本接口返回：{release_req}')
            all_logger.info('版本发布失败')
            raise LinuxCloudOTAError('版本发布失败')
        else:
            all_logger.info('版本发布成功')

    def create_upgrade_plan(self, is_forward):
        """
        创建升级计划,升级计划存在六种状态，每种状态进行不同处理，激活并审核计划
        :param is_forward:是否正向升级，True:正向升级，False:反向升级
        :return:
        """
        upgrade_plan_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/list?isOperation=false&planName=&targetVersion=&projectId={self.product_id}&page=1&size=50'
        upgrade_plan_req = requests.get(url=upgrade_plan_url, headers=self.authorization).json()
        if not upgrade_plan_req['success']:
            all_logger.info(f'请求查看当前升级计划接口返回:{upgrade_plan_req}')
        current_version = self.ati if is_forward else self.prev_ati
        target_version = self.prev_ati if is_forward else self.ati
        for i in upgrade_plan_req['result']['data']:
            if i['currentVersion'] == current_version and i['targetVersion'] == target_version and i['planType'] == 'FORMAL':
                plan_id = i['id']   # 获取升级计划的plan_id,用于后续配置策略激活计划等
                if i['status'] == "UN_ACTIVE":
                    all_logger.info('存在升级计划但未激活，配置升级策略后激活并通过审批')
                    expire_date_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time() + (86400 * 3)))
                    change_strategy_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/savePlanStrategy?planId={plan_id}&retryTimes=30&retryIntervalMins=0&timeoutMins=10&signalStrength=-1200&batteryCapacity=30&moduleFreeSpace=200&communityModuleLimit=0&expireDateStr={expire_date_str}'
                    change_req = requests.get(url=change_strategy_url, headers=self.authorization).json()
                    if not change_req['success']:
                        all_logger.info(f'请求修改升级策略接口返回：{change_req}')
                        raise LinuxCloudOTAError(f'请求修改升级策略失败，返回{change_req}')
                    activate_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/activePlan?planId={plan_id}'
                    activate_req = requests.get(url=activate_url, headers=self.authorization).json()
                    if not activate_req['success']:
                        all_logger.info(f'请求激活测试计划接口返回：{activate_req}')
                        raise LinuxCloudOTAError(f'请求激活测试计划失败，返回{activate_req}')
                    approve_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/approvePlan?planId={plan_id}&isPass=true&approveComment=cloud_ota版本测试'
                    approve_req = requests.get(url=approve_url, headers=self.authorization).json()
                    if not approve_req['success']:
                        all_logger.info(f'请求审批通过该升级计划返回:{approve_req}')
                        raise LinuxCloudOTAError(f'请求审批通过该升级计划失败，返回{approve_req}')
                    all_logger.info('该升级计划已通过审批')
                    return True
                elif i['status'] == "WAITING_APPROVE":
                    all_logger.info('存在升级计划且已激活，配置升级策略后去客户运营中心审核通过')
                    expire_date_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time() + (86400 * 3)))
                    change_strategy_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/savePlanStrategy?planId={plan_id}&retryTimes=30&retryIntervalMins=0&timeoutMins=10&signalStrength=-1200&batteryCapacity=30&moduleFreeSpace=200&communityModuleLimit=0&expireDateStr={expire_date_str}'
                    change_req = requests.get(url=change_strategy_url, headers=self.authorization).json()
                    if not change_req['success']:
                        all_logger.info(f'请求修改升级策略接口返回：{change_req}')
                        raise LinuxCloudOTAError(f'请求修改升级策略失败，返回{change_req}')
                    approve_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/approvePlan?planId={plan_id}&isPass=true&approveComment=cloud_ota版本测试'
                    approve_req = requests.get(url=approve_url, headers=self.authorization).json()
                    if not approve_req['success']:
                        all_logger.info(f'请求审批通过该升级计划返回:{approve_req}')
                        raise LinuxCloudOTAError(f'请求审批通过该升级计划失败，返回{approve_req}')
                    all_logger.info('该升级计划已通过审批')
                    return True
                elif i['status'] == "UPGRADING":
                    all_logger.info('存在升级计划且激活并审核通过，配置升级策略后直接发指令升级')
                    expire_date_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time() + (86400 * 3)))
                    change_strategy_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/savePlanStrategy?planId={plan_id}&retryTimes=30&retryIntervalMins=0&timeoutMins=10&signalStrength=-1200&batteryCapacity=30&moduleFreeSpace=200&communityModuleLimit=0&expireDateStr={expire_date_str}'
                    change_req = requests.get(url=change_strategy_url, headers=self.authorization).json()
                    if not change_req['success']:
                        all_logger.info(f'请求修改升级策略接口返回：{change_req}')
                        raise LinuxCloudOTAError(f'请求修改升级策略失败，返回{change_req}')
                    return True
                elif i['status'] == 'UPGRADE_COMPLETE':
                    all_logger.info('存在升级计划，且已升级完成，不可再用，需要重新建立升级计划并且激活，通过审批')
                    plan_id = self.add_upgrade_plan(True if is_forward else False)
                    approve_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/approvePlan?planId={plan_id}&isPass=true&approveComment=cloud_ota版本测试'
                    approve_req = requests.get(url=approve_url, headers=self.authorization).json()
                    if not approve_req['success']:
                        all_logger.info(f'请求审批通过该升级计划返回:{approve_req}')
                        raise LinuxCloudOTAError(f'请求审批通过该升级计划失败，返回{approve_req}')
                    all_logger.info('该升级计划已通过审批')
                    return True
                elif i['status'] == 'APPROVE_FAILED':
                    all_logger.info('存在升级计划审核不通过，直接删除，之后重新添加升级计划')
                    delete_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/delete?id={plan_id}'
                    delete_req = requests.get(url=delete_url, headers=self.authorization).json()
                    if not delete_req['success']:
                        all_logger.info(f'请求删除升级计划接口返回：{delete_req}')
                        raise LinuxCloudOTAError(f'请求删除升级计划失败，返回{delete_req}')
                    self.add_upgrade_plan(True if is_forward else False)
                    approve_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/approvePlan?planId={plan_id}&isPass=true&approveComment=cloud_ota版本测试'
                    approve_req = requests.get(url=approve_url, headers=self.authorization).json()
                    if not approve_req['success']:
                        all_logger.info(f'请求审批通过该升级计划返回:{approve_req}')
                        raise LinuxCloudOTAError(f'请求审批通过该升级计划失败，返回{approve_req}')
                    all_logger.info('该升级计划已通过审批')
                    return True
            else:
                if i['status'] == "UN_ACTIVE":
                    all_logger.info('存在无关升级计划，且未激活，删除')
                    delete_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/delete?id={i["id"]}'
                    delete_req = requests.get(url=delete_url, headers=self.authorization).json()
                    if not delete_req:
                        all_logger.info(f'请求删除升级计划接口返回：{delete_req}')
                        raise LinuxCloudOTAError(f'请求删除升级计划失败，返回{delete_req}')
        else:
            all_logger.info('当前未存在升级计划，创建升级计划并激活，审批通过')
            plan_id = self.add_upgrade_plan(True if is_forward else False)
            approve_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/approvePlan?planId={plan_id}&isPass=true&approveComment=cloud_ota版本测试'
            approve_req = requests.get(url=approve_url, headers=self.authorization).json()
            if not approve_req['success']:
                all_logger.info(f'请求审批通过该升级计划返回:{approve_req}')
                raise LinuxCloudOTAError(f'请求审批通过升级计划失败，返回{approve_req}')
            all_logger.info('该升级计划已通过审批')

    def add_upgrade_plan(self, is_forward):
        """
        添加升级计划并激活
        :param is_forward:创建正向或反向升级计划：True:创建正向升级计划，False:创建反向升级计划
        :return:
        """
        upgrade_plan_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/list?isOperation=false&planName=&targetVersion=&projectId={self.product_id}&page=1&size=50'
        all_logger.info('添加升级计划')
        current_version = self.ati if is_forward else self.prev_ati
        target_version = self.prev_ati if is_forward else self.ati

        # 第一步，发送请求遍历版本列表获取plan_id
        url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/version/listByPageForUser?page=1&size=10&projectId={self.product_id}'
        get_version_req = requests.get(url=url, headers=self.authorization).json()
        if not get_version_req['success']:
            all_logger.info(f'获取版本列表失败，返回{get_version_req}')
        get_id_val = get_version_req['result']['data']
        plan_id = ''
        for i in get_id_val:
            if i['currentVersion'] == current_version and i['targetVersion'] == target_version:
                plan_id = i['id']

        # 第二步，发送请求创建计划
        plan_name = 'forward_upgrade' if is_forward else 'reverse_upgrade'
        save_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/saveFormalPlan?planName={plan_name}&upgradeRange=ALL&upgradeType=DIFF_PACK&versionId={plan_id}&projectId={self.product_id}'
        save_req = requests.post(url=save_url, headers=self.authorization).json()
        if not save_req['success']:
            all_logger.info(f'请求创建升级计划失败，返回{save_req}')
        all_logger.info('已添加升级计划，配置升级策略后激活')

        # 第四步，配置升级策略
        upgrade_plan_req = requests.get(url=upgrade_plan_url, headers=self.authorization).json()
        if not upgrade_plan_req['success']:
            all_logger.info(upgrade_plan_req)
        plan_id = upgrade_plan_req['result']['data'][0]['id']
        expire_date = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time() + (86400 * 3)))
        change_strategy_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/savePlanStrategy?planId={plan_id}&retryTimes=30&retryIntervalMins=0&timeoutMins=10&signalStrength=-1200&batteryCapacity=30&moduleFreeSpace=200&communityModuleLimit=0&expireDateStr={expire_date}'
        change_req = requests.get(url=change_strategy_url, headers=self.authorization).json()
        if not change_req['success']:
            all_logger.info(f'请求配置升级策略失败，返回{change_req}')

        # 第五步，激活升级计划
        activate_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/activePlan?planId={plan_id}'
        activate_req = requests.get(url=activate_url, headers=self.authorization).json()
        if not activate_req['success']:
            all_logger.info(f'请求激活升级计划失败，返回{activate_req}')
            raise LinuxCloudOTAError(f'请求激活升级计划失败，返回{activate_req}')
        return plan_id

    def check_info_before_upgrade(self):
        """
        升级前模块信息查询
        :return:
        """
        self.at_handle.send_at('AT+QPRTPARA=1', 30)
        self.at_handle.send_at('AT+QCFG="ApRstLevel",0', 10)
        self.at_handle.send_at('AT+QCFG="ModemRstLevel",0', 10)
        self.at_handle.check_network()

    def prepare_before_upgrade(self):
        """
        升级前准备，主要用于添加PK,以及HTTP的URL
        :return:
        """
        # 首先请求产品列表内容，获取产品的projectKey和encryptKey
        product_url = f'https://iot-gateway.quectel.com/v2/cloudotaportal/project/list?page=1&rows=10&projectName=&userId={self.user_id}&cloudProjectId={self.project_id}'
        product_req = requests.get(url=product_url, headers=self.authorization).json()
        project_key = ''
        encrypt_key = ''
        if not product_req['success']:
            all_logger.info(f'请求产品列表页面现有产品接口返回：{product_req}')
        for i in product_req['result']['content']:
            if i['projectName'] == f'{self.module_model}_ota_upgrade':
                project_key = i['projectKey']
                encrypt_key = i['encryptKey']
        self.at_handle.send_at('AT+QFOTACFG="server","https://iot-gateway.quectel.com"', 10)
        self.at_handle.send_at('AT+QFOTACFG="tls",1', 10)
        self.at_handle.send_at(f'AT+QFOTACFG="PK","{encrypt_key}","{project_key}"', 10)
        value = self.at_handle.send_at('AT+QFOTACFG?', 10)
        set_value_list = ['"server","https://iot-gateway.quectel.com"', '"tls",1', f'"pk","{encrypt_key}","{project_key}"']
        for i in set_value_list:
            if i not in value:
                all_logger.info(f'AT+QFOTACFG?指令查询，存在设置{i}未生效，返回结果为{value}')
                raise LinuxCloudOTAError(f'AT+QFOTACFG?指令查询，存在设置{i}未生效，返回结果为{value}')
        else:
            all_logger.info('AT+QFOTACFG相关指令设置完成')

    def ota_upgrade(self):
        """
        发指令进行升级
        :return:
        """
        self.at_handle.send_at('AT+QFOTAUP=100', 10)
        error_urc_dict = {'+QFOTAEVT: "fota",2': '请求升级失败', '+QFOTAEVT: "fota",5': '升级包下载失败'}
        with serial.Serial(self.at_port, baudrate=115200, timeout=0) as __at_port:
            start_time = time.time()
            return_val_cache = ''
            while True:
                time.sleep(0.001)
                return_val = self.at_handle.readline(__at_port)
                if return_val != '':
                    return_val_cache += return_val
                if '+QFOTAEVT: "fota",3' in return_val:
                    all_logger.info('模块开始下载升级包')
                if '+QFOTAEVT: "fota",4' in return_val:
                    all_logger.info('升级包下载完成')
                if '+QFOTAEVT: "fota",6' in return_val:
                    all_logger.info('模块开始升级,等待掉口后上报升级URC')
                    start_time = time.time()
                    break
                for k, v in error_urc_dict.items():
                    if k in return_val_cache:
                        all_logger.info(v)
                        raise LinuxCloudOTAError(v)
                if time.time() - start_time > 50:
                    raise LinuxCloudOTAError('50S内未检测到升级包下载完成URC')
        if not self.driver_check.check_usb_driver_dis():
            raise LinuxCloudOTAError('升级包下载完成后模块未掉口')
        finish_time = time.time()
        self.driver_check.check_usb_driver()
        self.check_upgrade_urc()
        return finish_time - start_time

    def check_upgrade_urc(self):
        """
        检测模块掉口重新上口后升级URC
        :return:
        """
        return_val_cache = ''
        with serial.Serial(self.debug_port, baudrate=115200, timeout=0) as __at_port:
            start_time = time.time()
            while True:
                time.sleep(0.001)
                return_val = self.at_handle.readline(__at_port)
                return_val_cache += return_val
                if '+QIND: "FOTA","END",0' in return_val:
                    all_logger.info('cloud_ota升级成功,等待模块重启')
                    break
                if time.time() - start_time > 300:
                    all_logger.info('300S内模块未升级成功')
                    raise LinuxCloudOTAError('300S内模块未升级成功')
        self.driver_check.check_usb_driver_dis()
        self.driver_check.check_usb_driver()
        self.at_handle.check_urc()
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",7', timout=120)
        if '+QIND: "FOTA","START"' not in return_val_cache:
            all_logger.info('升级过程中未上报"FOTA","START"')
            raise LinuxCloudOTAError('升级过程中未上报"FOTA","START"')

    def check_module_info(self, is_forward, imei):
        """
        升级后检查版本号及IMEI号等
        :param is_forward: 是否正向升级，True:正向升级，False:反向升级
        :param imei: 模块imei号，默认是下发的imei号，也存在改变imei号后升级的情况
        :return:
        """
        revision = ''
        for _ in range(10):
            return_value = self.at_handle.send_at('ATI+CSUB', 0.6)
            revision = ''.join(re.findall(r'Revision: (.*)\r', return_value)).strip()
            sub_edition = ''.join(re.findall(r'SubEdition: (.*)\r', return_value)).strip()
            if is_forward is False:     # 反向升级后查询版本号，期望版本号为测试版本
                if revision == self.ati and sub_edition == self.sub:
                    break
            else:     # 正向升级后查询版本号，期望版本号不为测试版本
                if revision == self.prev_ati and sub_edition == self.prev_sub:
                    break
            time.sleep(1)
        else:
            all_logger.info(f'ATI查询的版本号和当前设置版本号不一致,当前ATI版本为{revision}，期望ATI版本为{self.prev_ati if is_forward else self.ati}')
            raise LinuxCloudOTAError(f'ATI查询的版本号和当前设置版本号不一致,当前ATI版本为{revision}，期望ATI版本为{self.prev_ati if is_forward else self.ati}')
        # 查号
        return_value = self.at_handle.send_at('AT+EGMR=0,7', 0.3)
        if 'OK' not in return_value or imei not in return_value:
            raise LinuxCloudOTAError("升级后IMEI号异常")

    def white_list(self, is_forward):
        """
        将模块IMEI号导入白名单
        :param is_forward: 判断正向还是反向升级
        :return:
        """
        self.at_handle.send_at(f'AT+EGMR=1,7,"{self.imei}"')
        upgrade_plan_list_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/list?isOperation=false&planName=&targetVersion=&projectId={self.product_id}&page=1&size=10'
        upgrade_plan_req = requests.get(url=upgrade_plan_list_url, headers=self.authorization).json()
        if not upgrade_plan_req['success']:
            all_logger.info(f'请求获取升级计划列表信息返回:{upgrade_plan_req}')
        plan_id = ''
        for i in upgrade_plan_req['result']['data']:
            cur_version = self.ati if is_forward else self.prev_ati
            tar_version = self.prev_ati if is_forward else self.ati
            if i['currentVersion'] == cur_version and i['targetVersion'] == tar_version and i['status'] == 'UPGRADING':
                plan_id = i['id']
                break
        url = r'https://iot-gateway.quectel.com/zuul/v2/cloudotaportal/planBwList/importBwList'
        self.create_excel()     # 创建excel表格写入imei号
        data = {'planId': plan_id, 'bwListType': 'WHITE_LIST'}
        file_path = os.path.join(os.getcwd(), "imeiImport.xlsx")
        files = {'file': open(file_path, 'rb')}
        add_white_list_req = requests.post(url=url, data=data, files=files, headers=self.authorization).json()
        if not add_white_list_req['success']:
            all_logger.info(add_white_list_req)
            raise LinuxCloudOTAError(f'添加模块IMEI号白名单失败，返回{add_white_list_req}')
        all_logger.info('添加模块IMEI号白名单成功')
        return plan_id

    def delete_white_list(self, plan_id):
        """
        删除白名单
        :param plan_id:白名单对应的id
        :return:
        """
        white_list_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/planBwList/page?page=1&size=10&planId={plan_id}&imei=&bwListType=WHITE_LIST'
        id_req = requests.get(url=white_list_url, headers=self.authorization).json()
        white_id = id_req['result']['data'][0]['id']
        delete_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/planBwList/batchDelete?ids={white_id}'
        delete_req = requests.post(url=delete_url, headers=self.authorization).json()
        if not delete_req['success']:
            all_logger.info(f'请求删除白名单imei号返回:{delete_req}')
            all_logger.info('删除白名单imei号失败')
            raise LinuxCloudOTAError('删除白名单imei号失败')

    def create_excel(self):
        """
        生成excel表格，填入IMEI号用于上传白名单
        :return:
        """
        path = os.path.join(os.getcwd(), "imeiImport.xlsx")
        if os.path.exists(path):
            all_logger.info('当前已存在imei导入表格')
            return True
        wb = Workbook()
        # 获取第一个sheet
        ws = wb.active
        # 将数据写入到指定的单元格
        ws['A1'] = 'IMEI'
        ws['A2'] = self.imei   # 写入IMEI号

        # 保存为imeiImport.xlsx
        wb.save("imeiImport.xlsx")

    def add_empty_product(self):
        """
        添加空产品，不含升级计划
        :return:
        """
        all_logger.info('添加空白产品，不含测试计划')
        module_id = self.get_module_id()    # 获取当前测试硬件型号在系统中的硬件ID
        add_product_url = f'https://iot-gateway.quectel.com/v2/cloudotaportal/project/save?projectName=empty_test&type=0&moduleId={module_id}&details=版本测试&isZh=true&qbProjectId={self.project_id}'
        add_product_req = requests.post(url=add_product_url, headers=self.authorization).json()
        if not add_product_req['success']:
            if add_product_req['msg'] == 'the product for the product name already exists!':
                all_logger.info('添加空白产品成功')
            else:
                all_logger.info(f'请求产品列表页添加产品接口返回：{add_product_req}')
                raise LinuxCloudOTAError(f'添加空白产品失败，返回{add_product_req}')
        product_url = f'https://iot-gateway.quectel.com/v2/cloudotaportal/project/list?page=1&rows=10&projectName=&userId={self.user_id}&cloudProjectId={self.project_id}'
        product_req = requests.get(url=product_url, headers=self.authorization).json()
        if not product_req['success']:
            all_logger.info(f'请求产品列表页面现有产品接口返回：{product_req}')
        cur_product_list = product_req['result']['content']
        for i in cur_product_list:
            if 'empty_test' in i['projectName']:
                return [i['projectKey'], i['encryptKey']]

    def set_empty_product(self, product_info):
        """
        修改PK为空产品，没有测试计划
        :param product_info: 产品信息，ID及密钥
        :return:
        """
        self.at_handle.send_at('AT+QPRTPARA=1', 10)
        self.at_handle.send_at(f'AT+QFOTACFG="PK","{product_info[1]}","{product_info[0]}"', 10)
        value = self.at_handle.send_at('AT+QFOTACFG?', 10)
        if product_info[1] not in value and product_info[0] not in value:
            all_logger.info('修改PK为空产品失败')
            raise LinuxCloudOTAError('修改PK为空产品失败')

    def set_error_product(self):
        """
        设置错误密钥
        :return:
        """
        self.at_handle.send_at('AT+QPRTPARA=1', 10)
        self.at_handle.send_at('AT+QFOTACFG="PK","123456","654321"', 10)
        value = self.at_handle.send_at('AT+QFOTACFG?', 10)
        if '123456' not in value and '654321' not in value:
            all_logger.info('修改PK为空产品失败')
            raise LinuxCloudOTAError('修改PK为错误产品失败')

    def ota_error_upgrade(self):
        """
        使用错误密钥升级
        :return:
        """
        self.at_handle.send_at('AT+QFOTAUP=100', 10)
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",2,10710', timout=120)

    def set_error_imei(self):
        """
        设置IMEI号为非白名单，并进行备份等
        :return:
        """
        self.at_handle.send_at('AT+EGMR=1,7,"123456781234567"')
        self.at_handle.send_at('AT+QPRTPARA=1', 10)
        self.at_handle.send_at('AT+QFOTACFG="tls",1', 10)

    def delay_upgrade(self):
        """
        设置延时后进行升级
        :return:
        """
        try:
            self.at_handle.send_at(f'AT+EGMR=1,7,"{self.imei}"', 10)
            self.prepare_before_upgrade()
            self.at_handle.send_at('AT+QFOTACFG="delay",2')
            delay_time = self.ota_upgrade()
            all_logger.info(f'上报差分包下载完成后{delay_time}S后掉口开始升级')
            if delay_time < 110:
                all_logger.info('设置延时升级两分钟，实际延时时间小于110S')
                raise LinuxCloudOTAError('设置延时升级两分钟，实际延时时间小于110S')
            elif delay_time > 130:
                all_logger.info('设置延时升级两分钟，实际延时时间大于130S')
                raise LinuxCloudOTAError('设置延时升级两分钟，实际延时时间大于130S')
        finally:
            self.at_handle.send_at('AT+QFOTACFG="delay",0')

    def low_power_upgrade(self):
        """
        低电量升级，预期无法升级
        :return:
        """
        self.at_handle.send_at('AT+QFOTAUP=1', 10)
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",2,10705', timout=120)

    def black_list(self, is_forward):
        """
        将模块IMEI号导入黑名单
        :param is_forward:根据当前版本决定修改哪个测试策略添加黑名单
        :return:
        """
        self.at_handle.send_at(f'AT+EGMR=1,7,"{self.imei}"')
        upgrade_plan_list_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/upgradePlan/list?isOperation=false&planName=&targetVersion=&projectId={self.product_id}&page=1&size=10'
        upgrade_plan_req = requests.get(url=upgrade_plan_list_url, headers=self.authorization).json()
        if not upgrade_plan_req['success']:
            all_logger.info(f'请求获取升级计划列表信息返回:{upgrade_plan_req}')
        plan_id = ''
        for i in upgrade_plan_req['result']['data']:
            cur_version = self.ati if is_forward else self.prev_ati
            tar_version = self.prev_ati if is_forward else self.ati
            if i['currentVersion'] == cur_version and i['targetVersion'] == tar_version and i['status'] == 'UPGRADING':
                plan_id = i['id']
                break
        url = r'https://iot-gateway.quectel.com/zuul/v2/cloudotaportal/planBwList/importBwList'
        self.create_excel()     # 创建excel表格写入imei号
        data = {'planId': plan_id, 'bwListType': 'BLACK_LIST'}
        file_path = os.path.join(os.getcwd(), "imeiImport.xlsx")
        files = {'file': open(file_path, 'rb')}
        try:
            add_white_list_req = requests.post(url=url, data=data, files=files, headers=self.authorization).json()
            if not add_white_list_req['success']:
                all_logger.info(f'请求添加黑名单返回:{add_white_list_req}')
                all_logger.info('添加黑名单失败')
                raise LinuxCloudOTAError('添加黑名单失败')
            all_logger.info('添加模块IMEI号黑名单成功')
            self.at_handle.send_at('AT+QFOTAUP=100', 10)
            self.at_handle.readline_keyword('+QFOTAEVT: "fota",2,10709', timout=120)
        finally:
            # 最后需要删除黑名单处理，首先查询id，之后根据id删除imei号
            black_list_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/planBwList/page?page=1&size=10&planId={plan_id}&imei=&bwListType=BLACK_LIST'
            id_req = requests.get(url=black_list_url, headers=self.authorization).json()
            black_id = id_req['result']['data'][0]['id']
            delete_url = fr'https://iot-gateway.quectel.com/v2/cloudotaportal/planBwList/batchDelete?ids={black_id}'
            delete_req = requests.post(url=delete_url, headers=self.authorization).json()
            if not delete_req['success']:
                all_logger.info(f'请求删除黑名单imei号返回:{delete_req}')
                all_logger.info('删除黑名单imei号失败')
                raise LinuxCloudOTAError('删除黑名单imei号失败')

    def vbat_before_download(self):
        """
        下载差分包未完成前断电，上报fota,4之后，未上报fota,6之前断电上电
        :return:
        """
        self.at_handle.send_at('AT+QFOTAUP=100', 10)
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",4', timout=120)
        self.gpio.set_vbat_high_level()
        self.driver_check.check_usb_driver_dis()
        time.sleep(3)
        self.gpio.set_vbat_low_level_and_pwk()
        self.driver_check.check_usb_driver()
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",6', timout=120)
        self.driver_check.check_usb_driver_dis()
        self.driver_check.check_usb_driver()
        self.at_handle.readline_keyword('+QIND: "FOTA","START"', '+QIND: "FOTA","END",0', timout=300)
        self.driver_check.check_usb_driver_dis()
        self.driver_check.check_usb_driver()
        self.at_handle.check_urc()
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",7', timout=120)
        self.check_module_info(self.version_flag, self.imei)

    def vbat_in_upgrade(self):
        """
        升级过程中断电
        :return:
        """
        self.at_handle.send_at('AT+QFOTAUP=100', 10)
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",4', '+QFOTAEVT: "fota",6', timout=120)
        self.driver_check.check_usb_driver_dis()
        self.driver_check.check_usb_driver()
        # 随机断电时间点定义:根据上报的updating进度，随机在上报次数1-10次时断电
        update_times = 0    # 升级上报计数器
        vbat_times = random.randint(1, 15)    # 升级上报计数器
        start_time = time.time()
        return_value_cache = ''
        with serial.Serial(self.at_port, baudrate=115200, timeout=0) as __at_port:
            while True:
                return_value = self.at_handle.readline(__at_port)
                return_value_cache += return_value
                if '+QIND: "FOTA","UPDATING"' in return_value:
                    update_times += 1
                if update_times == vbat_times:
                    all_logger.info('升级过程断电')
                    break
                if time.time() - start_time > 180 and '+QIND: "FOTA","END",0' not in return_value_cache:
                    all_logger.info('180S内未检测到升级上报结束URC')
                    raise LinuxCloudOTAError('180S内未检测到升级上报结束URC')
                if '+QIND: "FOTA","END",0' in return_value_cache and update_times == 0:
                    all_logger.info('已上报升级结束URC,但未收到升级进度上报URC')
                    raise LinuxCloudOTAError('已上报升级结束URC,但未收到升级进度上报URC')
                if '+QIND: "FOTA","END",0' in return_value_cache:
                    all_logger.info('已上报END 0，上报结束后断电')
                    break
        self.gpio.set_vbat_high_level()
        self.driver_check.check_usb_driver_dis()
        time.sleep(3)
        self.gpio.set_vbat_low_level_and_pwk()
        self.driver_check.check_usb_driver()
        self.at_handle.readline_keyword('+QIND: "FOTA","START"', '+QIND: "FOTA","END",0', timout=300)
        self.driver_check.check_usb_driver_dis()
        self.driver_check.check_usb_driver()
        self.at_handle.check_urc()
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",7', timout=120)
        self.check_module_info(self.version_flag, self.imei)

    def disconnect_network_upgrade(self):
        """
        下载差分包过程中断网
        :return:
        """
        self.at_handle.send_at('AT+QFOTAUP=100', 10)
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",3', timout=120)
        for i in range(10):
            value = self.at_handle.send_at('AT+QFLST')
            if 'otaupdate.zip' in value:
                all_logger.info('已检测到差分包开始下载，进行断网操作')
                self.at_handle.readline_keyword('+QFOTAEVT: "fota",5,10605', at_flag=True, at_cmd='AT+CFUN=0')
                all_logger.info('检测到差分包下载失败URC上报')
                break
            time.sleep(1)
        self.at_handle.send_at('AT+CFUN=1', 10)
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",3', '+QFOTAEVT: "fota",6', timout=300)
        self.driver_check.check_usb_driver_dis()
        self.driver_check.check_usb_driver()
        self.at_handle.readline_keyword('+QIND: "FOTA","START"', '+QIND: "FOTA","END",0', timout=300)
        self.driver_check.check_usb_driver_dis()
        self.driver_check.check_usb_driver()
        self.at_handle.check_urc()
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",7', timout=120)
        self.check_module_info(self.version_flag, self.imei)

    def vbat_in_download_upgrade(self):
        """
        下载差分包过程中断电三次继续升级
        :return:
        """
        self.at_handle.send_at('AT+QFOTAUP=100', 10)
        for i in range(3):
            self.at_handle.readline_keyword('+QFOTAEVT: "fota",3', timout=300)
            for j in range(10):
                value = self.at_handle.send_at('AT+QFLST')
                if 'otaupdate.zip' in value:
                    all_logger.info('已检测到差分包开始下载，进行断电操作')
                    break
                time.sleep(1)
            self.gpio.set_vbat_high_level()
            self.driver_check.check_usb_driver_dis()
            time.sleep(3)
            self.gpio.set_vbat_low_level_and_pwk()
            self.driver_check.check_usb_driver()
            self.at_handle.check_urc()
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",3', '+QFOTAEVT: "fota",6', timout=120)
        self.driver_check.check_usb_driver_dis()
        self.driver_check.check_usb_driver()
        self.at_handle.readline_keyword('+QIND: "FOTA","START"', '+QIND: "FOTA","END",0', timout=300)
        self.driver_check.check_usb_driver_dis()
        self.driver_check.check_usb_driver()
        self.at_handle.check_urc()
        self.at_handle.readline_keyword('+QFOTAEVT: "fota",7', timout=120)
        self.check_module_info(self.version_flag, self.imei)
